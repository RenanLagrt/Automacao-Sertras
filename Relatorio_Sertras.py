import os
import time
import subprocess
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from openpyxl.drawing.image import Image
from openpyxl.cell.cell import MergedCell
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def initialize_driver():
    return webdriver.Chrome(service=Service(ChromeDriverManager().install()))

def login_sertras(driver, email, senha):
    driver.get("https://gestaodeterceiros.sertras.com/escolha-um-contrato")
    driver.maximize_window()

    campo_email = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edtLoginInfo"]')))
    campo_email.send_keys(email)

    campo_senha = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edtLoginSenha"]')))
    campo_senha.send_keys(senha)

    botão_enter = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="btnLogin"]/div[2]')))
    botão_enter.click()

    fechar_janela = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="notificationPopup"]/div/div/div[1]/button/span')))
    fechar_janela.click()

def download_xml(driver):
    botão_relatório = WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/div/ul/li[9]/a/span[1]')))
    driver.execute_script("arguments[0].scrollIntoView();", botão_relatório)
    botão_relatório.click()

    botão_integração = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/div/ul/li[9]/ul/li[2]/a/span[1]')))
    driver.execute_script("arguments[0].scrollIntoView();", botão_integração)
    botão_integração.click()

    botão_integração_pessoas = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/div/ul/li[9]/ul/li[2]/ul/li[1]/a')))
    driver.execute_script("arguments[0].scrollIntoView();", botão_integração_pessoas)
    botão_integração_pessoas.click()

    marcar_todos = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="dashboard-v1"]/div[3]/div/div/div[2]/form/div[1]/div[1]/div/label/a[1]')))
    marcar_todos.click()

    botão_dowload = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="box-filter"]/button[3]')))
    botão_dowload.click()

def wait_for_download(diretorio_downloads):
    arquivos_iniciais = set(os.listdir(diretorio_downloads))

    while True:
        arquivos_atuais = set(os.listdir(diretorio_downloads))
        novos_arquivos = arquivos_atuais - arquivos_iniciais
        if novos_arquivos:
            arquivo_mais_recente = max(novos_arquivos, key=lambda f: os.path.getmtime(os.path.join(diretorio_downloads, f)))
            caminho_arquivo = os.path.join(diretorio_downloads, arquivo_mais_recente)
            if arquivo_mais_recente.endswith(".crdownload"):
                time.sleep(1)
            else:
                tamanho_anterior = -1
                while True:
                    if os.path.exists(caminho_arquivo):
                        tamanho_atual = os.path.getsize(caminho_arquivo)
                        if tamanho_atual == tamanho_anterior:
                            break
                        tamanho_anterior = tamanho_atual
                    time.sleep(1)
                break
        time.sleep(1)
    return caminho_arquivo

def ler_xml(caminho_arquivo):
    tree = ET.parse(caminho_arquivo)
    root = tree.getroot()

    dados = []
    colunas = []
    is_header = True

    for elem in root.iter():
        if elem.tag == '{urn:schemas-microsoft-com:office:spreadsheet}Row':
            if is_header:
                for cell in elem:
                    for child in cell:
                        if child.tag == '{urn:schemas-microsoft-com:office:spreadsheet}Data':
                            if child.text:
                                colunas.append(child.text.strip())
                is_header = False
            else:
                current_row = []
                for cell in elem:
                    for child in cell:
                        if child.tag == '{urn:schemas-microsoft-com:office:spreadsheet}Data':
                            if child.text:
                                current_row.append(child.text.strip())
                if current_row:
                    dados.append(current_row)

    return pd.DataFrame(dados, columns=colunas)

def tratar_tabela(df):
    colunas_para_remover = ["Contrato Terceiro", "Unidade", "Valor Preenchido", "Âmbito","Evento"]
    colunas_existentes = [col for col in colunas_para_remover if col in df.columns]

    if colunas_existentes:
        df = df.drop(columns=colunas_existentes)

    df = df.rename(columns={"Data da Última Análise": "Data Análise"})

    df = df.rename(columns=lambda x: x.upper())

    df["DOCUMENTO"] = df["DOCUMENTO"].replace({
                "CERTIFICADO OU REGISTRO DE CLASSE SUPERIOR E/OU TÉCNICO": "CERTIFICADO DE CLASSE",
                "CTPS OU RELATÓRIO DO E-SOCIAL": "CTPS",
                "DOCUMENTO DE IDENTIFICAÇÃO": "IDENTIFICAÇÃO",
                "FICHA DE ENTREGA DE EPI": "EPI",
                "CERTIFICADO NR 10": "NR10",
                "CERTIFICADO NR 11": "NR11",
                "CERTIFICADO NR 12": "NR12",
                "CERTIFICADO NR 33": "NR33",
                "CERTIFICADO NR 35": "NR35"
            })

    return df 

def ajustar_largura_colunas(ws):
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter 
        max_length = max((len(str(cell.value)) for cell in col_cells if cell.value), default=0)
        ws.column_dimensions[col_letter].width = max_length + 2

def personalizar_excel(caminho_saida):
    wb = load_workbook(caminho_saida)
    ws = wb.active

    alinhamento_central = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fundo_preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fonte_branca = Font(color="FFFFFF", bold=True, size=12)

    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws.row_dimensions[1].height = 30

    for cell in ws[1]:  
        if isinstance(cell, MergedCell):  
            continue
        cell.fill = fundo_preto
        cell.font = fonte_branca
        cell.alignment = alinhamento_central
        cell.border = borda

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 23
        for cell in row:
            if isinstance(cell, MergedCell):  
                continue
            cell.border = borda 
            cell.alignment = alinhamento_central

    ajustar_largura_colunas(ws)

    ws.freeze_panes = "B1"
    ws.auto_filter.ref = "A1:N{}".format(ws.max_row)

    wb.save(caminho_saida)

def criar_excel(df):
    data_atual = datetime.now().strftime("%d-%m-%Y")
    caminho_saida = (f"RELATÓRIO_SERTRAS {data_atual}.xlsx")

    df = tratar_tabela(df)

    df.to_excel(caminho_saida, index=False)

    personalizar_excel(caminho_saida)

    subprocess.run(["cmd", "/c", "start", "", caminho_saida], shell=True)

def main(email,senha):
    driver = initialize_driver()
    login_sertras(driver,email,senha)
    download_xml(driver)

    diretorio_downloads = diretorio_downloads = os.path.expanduser("~/Downloads")
    caminho_arquivo = wait_for_download(diretorio_downloads)

    df = ler_xml(caminho_arquivo)

    criar_excel(df)




load_dotenv()

email = os.getenv("EMAIL")
senha = os.getenv("SENHA")

main(email, senha)



