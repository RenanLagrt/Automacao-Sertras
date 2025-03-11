import os
import re
import io
import sys
import time
import psutil 
import subprocess
import pytesseract
import xml.etree.ElementTree as ET
import pandas as pd
import streamlit as st
from pdf2image import convert_from_path
from datetime import datetime
from dotenv import load_dotenv
from selenium import webdriver
from openpyxl import load_workbook
from itertools import zip_longest
from datetime import datetime, timedelta
from openpyxl.cell.cell import MergedCell
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


poppler_path = os.path.join(os.path.expanduser("~"),"Downloads","Release-24.08.0-0","poppler-24.08.0","Library","bin")
tesseract_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = tesseract_path


class AutomaçãoSertras():

    def __init__(self, email, senha):
        self.email = email
        self.senha = senha
        self.driver = None

    def initialize_driver(self):
        return webdriver.Chrome(service=Service(ChromeDriverManager().install())) 

    def login_sertras(self):
        self.driver.get("https://gestaodeterceiros.sertras.com/escolha-um-contrato")
        self.driver.maximize_window()

        campo_email = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edtLoginInfo"]')))
        campo_email.send_keys(self.email)

        campo_senha = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="edtLoginSenha"]')))
        campo_senha.send_keys(self.senha)

        botão_enter = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="btnLogin"]/div[2]')))
        botão_enter.click()

        fechar_janela = WebDriverWait(self.driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="notificationPopup"]/div/div/div[1]/button/span')))
        fechar_janela.click()

    def download_xml(self):
        botão_relatório = WebDriverWait(self.driver,10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/div/ul/li[9]/a/span[1]')))
        self.driver.execute_script("arguments[0].scrollIntoView();", botão_relatório)
        botão_relatório.click()

        botão_integração = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/div/ul/li[9]/ul/li[2]/a/span[1]')))
        self.driver.execute_script("arguments[0].scrollIntoView();", botão_integração)
        botão_integração.click()

        botão_integração_pessoas = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="sidebar-menu"]/div/ul/li[9]/ul/li[2]/ul/li[1]/a')))
        self.driver.execute_script("arguments[0].scrollIntoView();", botão_integração_pessoas)
        botão_integração_pessoas.click()

        marcar_todos = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="dashboard-v1"]/div[3]/div/div/div[2]/form/div[1]/div[1]/div/label/a[1]')))
        marcar_todos.click()

        botão_dowload = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="box-filter"]/button[3]')))
        botão_dowload.click()

    @staticmethod
    def wait_for_download(diretorio_downloads, timeout=45):
        arquivos_iniciais = set(os.listdir(diretorio_downloads))
        tempo_inicio = time.time()

        while time.time() - tempo_inicio < timeout:
            time.sleep(0.5)  # Pequeno atraso para evitar chamadas excessivas
            arquivos_atuais = set(os.listdir(diretorio_downloads))
            novos_arquivos = arquivos_atuais - arquivos_iniciais

            for arquivo in novos_arquivos:
                caminho_arquivo = os.path.join(diretorio_downloads, arquivo)

                if not arquivo.endswith(".crdownload") and os.path.isfile(caminho_arquivo):
                    tamanho_anterior = -1
                    while time.time() - tempo_inicio < timeout:
                        time.sleep(0.5)  
                        if os.path.exists(caminho_arquivo):
                            tamanho_atual = os.path.getsize(caminho_arquivo)
                            if tamanho_atual > 0 and tamanho_atual == tamanho_anterior:
                                return caminho_arquivo
                            tamanho_anterior = tamanho_atual

        raise TimeoutError("Erro: Tempo limite atingido para download.")

    @staticmethod
    def ler_xml(caminho_arquivo):
        try:
            tree = ET.parse(caminho_arquivo)
            root = tree.getroot()

        except ET.ParseError:
            raise ValueError("Erro ao processar XML. Verifique se o arquivo está correto.")

        namespace = '{urn:schemas-microsoft-com:office:spreadsheet}'
        dados, colunas = [], []
        is_header = True

        for row in root.iter(f'{namespace}Row'):
            linha = [
                cell[0].text.strip() if len(cell) > 0 and cell[0].tag == f'{namespace}Data' and cell[0].text else ''
                for cell in row
            ]
            if is_header:
                colunas = linha
                is_header = False
            else:
                if any(linha):  # Evita adicionar linhas vazias
                    dados.append(linha)

        return pd.DataFrame(dados, columns=colunas) 

    @staticmethod
    def tratar_tabela(tabela_sertras):
        colunas_para_remover = ["Contrato Terceiro", "Unidade", "Valor Preenchido", "Âmbito","Evento"]
        colunas_existentes = [col for col in colunas_para_remover if col in tabela_sertras.columns]

        if colunas_existentes:
            tabela_sertras = tabela_sertras.drop(columns=colunas_existentes)

        tabela_sertras = tabela_sertras.rename(columns={"Data da Última Análise": "Data Análise"})

        tabela_sertras = tabela_sertras.rename(columns=lambda x: x.upper())

        tabela_sertras["DOCUMENTO"] = tabela_sertras["DOCUMENTO"].replace({
                    "CERTIFICADO OU REGISTRO DE CLASSE SUPERIOR E/OU TÉCNICO": "CERTIFICADO DE CLASSE",
                    "CTPS OU RELATÓRIO DO E-SOCIAL": "CTPS",
                    "DOCUMENTO DE IDENTIFICAÇÃO": "IDENTIFICAÇÃO",
                    "FICHA DE ENTREGA DE EPI": "EPI",
                    "CERTIFICADO NR 10": "NR10",
                    "CERTIFICADO NR 11": "NR11",
                    "CERTIFICADO NR 12": "NR12",
                    "CERTIFICADO NR 33": "NR33",
                    "CERTIFICADO NR 35": "NR35"
                })

        return tabela_sertras 

    @staticmethod
    def ajustar_largura_colunas(ws):
        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter 
            max_length = max((len(str(cell.value)) for cell in col_cells if cell.value), default=0)
            ws.column_dimensions[col_letter].width = max_length + 2

    def personalizar_excel(self, caminho_saida):
        wb = load_workbook(caminho_saida)
        ws = wb.active

        alinhamento_central = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fundo_preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        fonte_branca = Font(color="FFFFFF", bold=True, size=12)

        borda = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        ws.row_dimensions[1].height = 30

        for cell in ws[1]:  
            if isinstance(cell, MergedCell):  
                continue
            cell.fill = fundo_preto
            cell.font = fonte_branca
            cell.alignment = alinhamento_central
            cell.border = borda

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 23
            for cell in row:
                if isinstance(cell, MergedCell):  
                    continue
                cell.border = borda 
                cell.alignment = alinhamento_central

        self.ajustar_largura_colunas(ws)

        ws.freeze_panes = "B1"
        ws.auto_filter.ref = "A1:N{}".format(ws.max_row)

        wb.save(caminho_saida)

    def BaixarRelatório(self):
        self.driver = self.initialize_driver()
        self.login_sertras()
        self.download_xml()

        diretorio_downloads = os.path.expanduser("~/Downloads")
        caminho_arquivo = self.wait_for_download(diretorio_downloads)

        tabela_sertras = self.ler_xml(caminho_arquivo)

        tabela_sertras = self.tratar_tabela(tabela_sertras)

        return tabela_sertras

    def interacao_interface_recursos(self):
        botão_recursos = WebDriverWait(self.driver,10).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="sidebar-menu"]/div/ul/li[8]/a/span[1]')))
        self.driver.execute_script("arguments[0].scrollIntoView();", botão_recursos)
        botão_recursos.click()

        botão_recursos_pessoas = WebDriverWait(self.driver,10).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="sidebar-menu"]/div/ul/li[8]/ul/li[1]/a')))
        self.driver.execute_script("arguments[0].scrollIntoView();", botão_recursos_pessoas)
        botão_recursos_pessoas.click()

    def interacao_interface_envio(self, nome):
        campo_nome = WebDriverWait(self.driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="filtro_nome"]'))) 
        campo_nome.clear()
        campo_nome.send_keys(nome)

        botão_filtrar_nome = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH,  '//*[@id="dashboard-v1"]/div[4]/div/div/div[2]/form/div[6]/button[1]')))
        botão_filtrar_nome.click()

        for tentativa in range(4):
            try:
                botao_eventos = WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="data-tables2"]/tbody/tr/td[9]/a')))
                botao_eventos.click()
                break

            except StaleElementReferenceException:
                print(f"Tentativa {tentativa+1}/5 falhou: elemento foi recriado. Tentando novamente...")

        abas = self.driver.window_handles
        self.driver.switch_to.window(abas[-1])

        # Bloco para garantir o envio de documentos de funcionários demitidos, visto que o xpath do botão de documentação desses são alterados
        try:   
            botão_documentação = WebDriverWait(self.driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="data-tables2"]/tbody/tr[4]/td[4]/ul/li/a')))

        except TimeoutException:
            try:
                botão_documentação = WebDriverWait(self.driver, 2).until(EC.presence_of_element_located((By.XPATH, '//*[@id="data-tables2"]/tbody/tr[5]/td[4]/ul/li/a')))

            except TimeoutException:
                print(f"Nenhum botão de documentação encontrado para {nome}")
                st.error(f"Nenhum botão de documentação encontrado para {nome}")
                return

        self.driver.execute_script("arguments[0].scrollIntoView();", botão_documentação)
        botão_documentação.click()

        abas = self.driver.window_handles
        self.driver.switch_to.window(abas[-1])

    @staticmethod
    def extrair_texto_ocr(caminho_arquivo, poppler_path):
        paginas_imagem = convert_from_path(caminho_arquivo, poppler_path=poppler_path)
        return " ".join(pytesseract.image_to_string(pagina) for pagina in paginas_imagem)

    @staticmethod
    def extrair_datas(texto, padrao):
        return re.findall(padrao, texto)

    @staticmethod
    def calcular_vencimento(data_str, anos=1):
        data_obj = datetime.strptime(data_str, "%d/%m/%Y")
        return (data_obj.replace(year=data_obj.year + anos)).strftime("%d/%m/%Y")

    def ler_aso(self,caminho_arquivo, poppler_path):
        texto = self.extrair_texto_ocr(caminho_arquivo, poppler_path)
        datas = self.extrair_datas(texto, r'\b\d{2}/\d{2}/\d{4}\b')
        if len(datas) > 1:
            return datas[1], self.calcular_vencimento(datas[1])  
        return None, None  

    def ler_epi(self,caminho_arquivo, poppler_path):
        texto = self.extrair_texto_ocr(caminho_arquivo, poppler_path)
        datas = self.extrair_datas(texto, r'\b\d{2}/\d{2}/\d{2}\b')
        if datas:
            data = datas[-1]
            if len(data.split('/')[2]) == 2:
                data = data[:6] + '20' + data[6:]
            return data, self.calcular_vencimento(data)
        return None, None

    def ler_Nrs(self,caminho_arquivo, poppler_path, documento):
        texto = self.extrair_texto_ocr(caminho_arquivo, poppler_path)
        datas = self.extrair_datas(texto, r'(\d{1,2}/\d{1,2}/\d{4})|(\d{1,2} de [a-zà-ú]+ de \d{4})')

        meses = {'janeiro': '01', 'fevereiro': '02', 'março': '03', 'abril': '04',
                'maio': '05', 'junho': '06', 'julho': '07', 'agosto': '08',
                'setembro': '09', 'outubro': '10', 'novembro': '11', 'dezembro': '12'}
        
        for data in datas:
            data = data[0] or data[1]
            if 'de' in data:
                partes = data.split(' de ')
                if len(partes) == 3:
                    data = f"{partes[0].zfill(2)}/{meses.get(partes[1])}/{partes[2]}"
            
            anos = 2 if documento in ["NR35", "NR10"] else 1
            return data, self.calcular_vencimento(data, anos)  

        return None, None

    def extrair_vencimento(self,caminho_arquivo, poppler_path, documento):
        if documento == "ASO":
            return self.ler_aso(caminho_arquivo, poppler_path)
        
        elif documento == "EPI":
            return self.ler_epi(caminho_arquivo, poppler_path)
        
        else:
            return self.ler_Nrs(caminho_arquivo, poppler_path, documento)

    @staticmethod
    def obter_data_modificacao(caminho_arquivo):
        return datetime.fromtimestamp(os.path.getmtime(caminho_arquivo))

    def verificar_atualizacao(self,status, data_analise, data_envio, caminho_arquivo):
        data_modificacao = self.obter_data_modificacao(caminho_arquivo)

        if status == "Pendente Correção":
            data_analise = datetime.strptime(data_analise, "%d/%m/%Y")
            return data_modificacao > data_analise, data_modificacao.strftime("%d/%m/%Y %H:%M")
        
        else:
            data_envio = datetime.strptime(data_envio, "%d/%m/%Y %H:%M")
            return data_modificacao > data_envio, data_modificacao.strftime("%d/%m/%Y %H:%M")

    def enviar_documento(self, documentos_validos, mapeamento_documentos, mapeamento_datas, vencimentos_enviados, documentos_enviados):
        for arquivo, documento, caminho_arquivo, data_vencimento, função in documentos_validos:
            if documento in mapeamento_datas:
                xpath_data = mapeamento_datas[documento]
                campo_data = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_data)))
                self.driver.execute_script("arguments[0].scrollIntoView();", campo_data)
                campo_data.clear()
                campo_data.send_keys(data_vencimento)
                vencimentos_enviados.append(data_vencimento)

            if documento in mapeamento_documentos:
                xpath_documento = mapeamento_documentos[documento]
                botao_upload = WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_documento)))
                self.driver.execute_script("arguments[0].scrollIntoView();", botao_upload)
                botao_upload.send_keys(caminho_arquivo)
                documentos_enviados.append(arquivo)

            time.sleep(1)

            botao_envio = WebDriverWait(self.driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="btnFuncaoRequisitoValores"]')))
            self.driver.execute_script("arguments[0].scrollIntoView();", botao_envio)
            botao_envio.click()

        time.sleep(2)
        abas = self.driver.window_handles
        self.driver.switch_to.window(abas[-1])
        self.driver.close()
        self.driver.switch_to.window(abas[-2])
        self.driver.close()
        self.driver.switch_to.window(abas[0])

    def run_complete_automation(self, documentos_rh, documentos_QSMS, diretorio_base_rh, diretorio_base_qsms, mapeamento_para_documentos, mapeamento_para_datas):
        tabela_sertras = self.BaixarRelatório()

        tabela_sertras = self.tratar_tabela(tabela_sertras)
        tabela_sertras = tabela_sertras.drop(["COMENTÁRIO ANALISTA", "PRAZO SLA"], axis=1)

        Status = ["Pendente", "Pendente Correção", "Vencido"]
        tabela_sertras = tabela_sertras[tabela_sertras["STATUS"].isin(Status)]
        self.interacao_interface_recursos()

        documentos_enviados = []
        erro_envio = []
        documentos_não_encontrados = []
        documentos_encontrados = []
        documentos_atualizados = []
        documentos_nao_atualizados = []
        datas_extraidas = []
        datas_modificacao = []
        vencimentos_projetados = []
        vencimentos_enviados = []


        NRs = ["NR10", "NR11", "NR12", "NR33", "NR35"]
        
        for nome, grupo in tabela_sertras.groupby("NOME"):
            documentos_validos = []

            for _, linha in grupo.iterrows():
                status, documento, funcao = linha["STATUS"], linha["DOCUMENTO"], linha["FUNÇÃO"]
                caminho_base = diretorio_base_rh if documento in documentos_rh else diretorio_base_qsms
                arquivo = f"{documento} - {nome}"
                
                caminho_arquivo = os.path.join(os.path.expanduser("~"), caminho_base, nome, f"{arquivo}.pdf")

                if not os.path.exists(caminho_arquivo):
                    documentos_não_encontrados.append(arquivo)
                    continue

                documentos_encontrados.append(arquivo)

                if status in ["Pendente Correção","Vencido"]:
                    atualizado, data_modificacao = self.verificar_atualizacao(status, linha["DATA ANÁLISE"], linha["DATA ENVIO"], caminho_arquivo)
                    datas_modificacao.append(data_modificacao)

                    if not atualizado:
                        documentos_nao_atualizados.append(arquivo)
                        continue
                    
                    else:
                        documentos_atualizados.append(arquivo)

                else:
                    datas_modificacao.append("N/A")

                data_extraida, data_vencimento = self.extrair_vencimento(caminho_arquivo, poppler_path, documento)
                
                if not data_vencimento:
                    erro_envio.append(arquivo)
                    continue

                if isinstance(data_vencimento, (list, tuple)):
                    data_vencimento = data_vencimento[0] if data_vencimento else None

                try:
                    data_vencimento = datetime.strptime(data_vencimento, "%d/%m/%Y")

                except (ValueError, TypeError):
                    erro_envio.append(arquivo)
                    continue

                if status == "Pendente Correção":
                    data_vencimento += timedelta(days=1) 

                data_vencimento = data_vencimento.strftime('%d/%m/%Y')
                datas_extraidas.append(data_extraida)
                vencimentos_projetados.append(data_vencimento)

                documentos_validos.append((arquivo, documento, caminho_arquivo, data_vencimento, funcao))

            if documentos_validos:
                self.interacao_interface_envio(nome)
            
                self.enviar_documento(documentos_validos, 
                                mapeamento_para_documentos.get(funcao, mapeamento_para_documentos["OUTRAS"]), 
                                mapeamento_para_datas.get(funcao, mapeamento_para_datas["OUTRAS"]), 
                                vencimentos_enviados, documentos_enviados)
    
        self.driver.quit()

        return tabela_sertras, documentos_não_encontrados, documentos_encontrados, documentos_enviados, datas_extraidas, vencimentos_projetados, vencimentos_enviados, erro_envio, documentos_atualizados, documentos_nao_atualizados, datas_modificacao  


# ---------------------------|-----------------------------------|-----------------------------------|----------------------------------|--------------------------------------|------

class RelatórioSertras(AutomaçãoSertras):

    def __init__(self, email, senha):
        super().__init__(email, senha)

    def criar_excel(self,tabela_sertras):
        data_atual = datetime.now().strftime("%d-%m-%Y")
        caminho_saida = (f"RELATÓRIO_SERTRAS {data_atual}.xlsx")

        tabela_sertras.to_excel(caminho_saida, index=False)

        self.personalizar_excel(caminho_saida)

        subprocess.run(["cmd", "/c", "start", "", caminho_saida], shell=True)

    def GerarRelatório(self):
        tabela_sertras = self.BaixarRelatório()

        tabela_sertras = self.tratar_tabela(tabela_sertras)

        self.criar_excel(tabela_sertras)

# ----------------------------|----------------------------------|--------------------------------------|-------------------------------------|-----------------------

class Envio_Sertras(AutomaçãoSertras):

    def __init__(self, email, senha):
        super().__init__(email, senha)

    def run_automation(self, documentos_rh, documentos_QSMS, diretorio_base_rh, diretorio_base_qsms, mapeamento_para_documentos, mapeamento_para_datas):
        data_atual = datetime.now().strftime("%d-%m-%Y")
        tabela_sertras = f"RELATÓRIO_SERTRAS {data_atual}.xlsx"
        
        tabela_sertras = pd.read_excel(tabela_sertras)

        tabela_sertras = tabela_sertras.drop(["COMENTÁRIO ANALISTA", "PRAZO SLA"], axis=1)

        Status = ["Pendente", "Pendente Correção", "Vencido"]
        tabela_sertras = tabela_sertras[tabela_sertras["STATUS"].isin(Status)]

        self.driver = self.initialize_driver()
        self.login_sertras()
        self.interacao_interface_recursos()

        documentos_enviados = []
        erro_envio = []
        documentos_não_encontrados = []
        documentos_encontrados = []
        documentos_atualizados = []
        documentos_nao_atualizados = []
        datas_extraidas = []
        datas_modificacao = []
        vencimentos_projetados = []
        vencimentos_enviados = []

        NRs = ["NR10", "NR11", "NR12", "NR33", "NR35"]
        
        for nome, grupo in tabela_sertras.groupby("NOME"):
            documentos_validos = []

            for _, linha in grupo.iterrows():
                status, documento, funcao = linha["STATUS"], linha["DOCUMENTO"], linha["FUNÇÃO"]
                caminho_base = diretorio_base_rh if documento in documentos_rh else diretorio_base_qsms
                arquivo = f"{documento} - {nome}"
                
                caminho_arquivo = os.path.join(os.path.expanduser("~"), caminho_base, nome, f"{arquivo}.pdf")

                if not os.path.exists(caminho_arquivo):
                    documentos_não_encontrados.append(arquivo)
                    continue

                documentos_encontrados.append(arquivo)

                if status in ["Pendente Correção","Vencido"]:
                    atualizado, data_modificacao = self.verificar_atualizacao(status, linha["DATA ANÁLISE"], linha["DATA ENVIO"], caminho_arquivo)
                    datas_modificacao.append(data_modificacao)

                    if not atualizado:
                        documentos_nao_atualizados.append(arquivo)
                        continue
                    
                    else:
                        documentos_atualizados.append(arquivo)

                else:
                    datas_modificacao.append("N/A")

                data_extraida, data_vencimento = self.extrair_vencimento(caminho_arquivo, poppler_path, documento)
                
                if not data_vencimento:
                    erro_envio.append(arquivo)
                    continue

                if isinstance(data_vencimento, (list, tuple)):
                    data_vencimento = data_vencimento[0] if data_vencimento else None

                try:
                    data_vencimento = datetime.strptime(data_vencimento, "%d/%m/%Y")

                except (ValueError, TypeError):
                    erro_envio.append(arquivo)
                    continue

                if status == "Pendente Correção":
                    data_vencimento += timedelta(days=1) 

                data_vencimento = data_vencimento.strftime('%d/%m/%Y')
                datas_extraidas.append(data_extraida)
                vencimentos_projetados.append(data_vencimento)

                documentos_validos.append((arquivo, documento, caminho_arquivo, data_vencimento, funcao))

            if documentos_validos:
                self.interacao_interface_envio(nome)
            
                self.enviar_documento(documentos_validos, 
                                mapeamento_para_documentos.get(funcao, mapeamento_para_documentos["OUTRAS"]), 
                                mapeamento_para_datas.get(funcao, mapeamento_para_datas["OUTRAS"]), 
                                vencimentos_enviados, documentos_enviados)
    
        self.driver.quit()

        return tabela_sertras, documentos_não_encontrados, documentos_encontrados, documentos_enviados, datas_extraidas, vencimentos_projetados, vencimentos_enviados, erro_envio, documentos_atualizados, documentos_nao_atualizados, datas_modificacao

# ----------------------------|-----------------------------------|-------------------------------------|------------------------------------|-------------------------------

class DocumentaçãoFuncionários():

    def __init__(self, contratos, diretorios_base, diretorios_dados, caminho_logo, documentos_por_função):
        self.contratos = contratos
        self.diretorios_base = diretorios_base
        self.diretorios_dados = diretorios_dados
        self.caminho_logo = caminho_logo
        self.documentos_por_função = documentos_por_função
    












load_dotenv()

email = os.getenv("EMAIL")
senha = os.getenv("SENHA")

data_atual = datetime.now().strftime("%d-%m-%Y") 

documentos_rh = ["IDENTIFICAÇÃO",
                "CTPS",
                "CONTRATO DE TRABALHO",
                "FICHA DE REGISTRO",
                "CNH"]
documentos_QSMS = ["ASO", 
                "EPI", 
                "NR10", 
                "NR11", 
                "NR12", 
                "NR33", 
                "NR35", 
                "CERTIFICADO DE CLASSE"]

diretorio_base_rh = os.path.join(os.path.expanduser("~"), "CONSORCIO CONCREJATOEFFICO LOTE 1", "Central de Arquivos - SERTRAS ARQUIVO PESSOAL")
diretorio_base_qsms =  os.path.join(os.path.expanduser("~"), "CONSORCIO CONCREJATOEFFICO LOTE 1", "Central de Arquivos - QSMS", "000 ATUAL - OBRA 186 - INHAÚMA", "Documentação Funcionários")  

mapeamento_para_documentos = {
        "ELETRICISTA DE REPARO DE REDE DE SANEAMENTO": {
            "IDENTIFICAÇÃO": '//*[@id="edtRequisito_Valor_1"]',
            "CTPS": '//*[@id="edtRequisito_Valor_2"]',
            "FICHA DE REGISTRO": '//*[@id="edtRequisito_Valor_3"]',
            "CONTRATO DE TRABALHO": '//*[@id="edtRequisito_Valor_4"]',
            "ASO": '//*[@id="edtRequisito_Valor_6"]',
            "EPI": '//*[@id="edtRequisito_Valor_8"]',
            "CERTIFICADO DE CLASSE": '//*[@id="edtRequisito_Valor_9"]',
            "NR10": '//*[@id="edtRequisito_Valor_11"]',
            "NR33": '//*[@id="edtRequisito_Valor_13"]',
            "NR35": '//*[@id="edtRequisito_Valor_15"]',
        },
        "OPERADOR DE ESCAVADEIRA": {
            "IDENTIFICAÇÃO": '//*[@id="edtRequisito_Valor_1"]',
            "CTPS": '//*[@id="edtRequisito_Valor_2"]',
            "FICHA DE REGISTRO": '//*[@id="edtRequisito_Valor_3"]',
            "CONTRATO DE TRABALHO": '//*[@id="edtRequisito_Valor_4"]',
            "ASO": '//*[@id="edtRequisito_Valor_6"]',
            "EPI": '//*[@id="edtRequisito_Valor_8"]',
            "CNH": '//*[@id="edtRequisito_Valor_10"]',
            "NR11": '//*[@id="edtRequisito_Valor_12"]',
        },
        "OUTRAS": {
            "IDENTIFICAÇÃO": '//*[@id="edtRequisito_Valor_1"]',
            "CTPS": '//*[@id="edtRequisito_Valor_2"]',
            "FICHA DE REGISTRO": '//*[@id="edtRequisito_Valor_3"]',
            "CONTRATO DE TRABALHO": '//*[@id="edtRequisito_Valor_4"]',
            "ASO": '//*[@id="edtRequisito_Valor_6"]',
            "EPI": '//*[@id="edtRequisito_Valor_8"]',
            "CERTIFICADO DE CLASSE": '//*[@id="edtRequisito_Valor_9"]',
            "NR10": '//*[@id="edtRequisito_Valor_11"]',
            "NR12": '//*[@id="edtRequisito_Valor_10"]',
            "NR33": '//*[@id="edtRequisito_Valor_12"]',
            "NR35": '//*[@id="edtRequisito_Valor_14"]',
        },
    }

mapeamento_para_datas = {
                "ELETRICISTA DE REPARO DE REDE DE SANEAMENTO" : {
                "IDENTIFICAÇÃO" : '//*[@id="edtRequisito_Valor_0"]',
                "ASO" : '//*[@id="edtRequisito_Valor_5"]',
                "EPI" : '//*[@id="edtRequisito_Valor_7"]',
                "NR10" : '//*[@id="edtRequisito_Valor_10"]',
                "NR33" : '//*[@id="edtRequisito_Valor_12"]',
                "NR35" : '//*[@id="edtRequisito_Valor_14"]',
            },
            "OPERADOR DE ESCAVADEIRA" : {
                "IDENTIFICAÇÃO" : '//*[@id="edtRequisito_Valor_0"]',
                "ASO" : '//*[@id="edtRequisito_Valor_5"]',
                "EPI" : '//*[@id="edtRequisito_Valor_7"]',
                "CNH" : '//*[@id="edtRequisito_Valor_9"]',
                "NR11" : '//*[@id="edtRequisito_Valor_11"]',
            },
            "OUTRAS" : {  
                "IDENTIFICAÇÃO" : '//*[@id="edtRequisito_Valor_0"]',
                "ASO" : '//*[@id="edtRequisito_Valor_5"]',
                "EPI" : '//*[@id="edtRequisito_Valor_7"]',
                "NR11" : '//*[@id="edtRequisito_Valor_11"]',
                "NR12" : '//*[@id="edtRequisito_Valor_9"]',
                "NR33" : '//*[@id="edtRequisito_Valor_11"]',
                "NR35" : '//*[@id="edtRequisito_Valor_13"]',
    },
}







automacao = Envio_Sertras(email, senha)


st.set_page_config(layout="wide")

if "executado" not in st.session_state:
    st.session_state["executado"] = False  

if "dados_processados" not in st.session_state:
    st.session_state["dados_processados"] = None

# Exibição do cabeçalho
logo_concrejato = os.path.join(os.path.expanduser("~"), "CONSORCIO CONCREJATOEFFICO LOTE 1", 
                            "Central de Arquivos - QSMS", "000 ATUAL - OBRA 201 - SÃO GONÇALO", 
                            "LOGO CONCREJATO.png")

logo_consorcio = os.path.join(os.path.expanduser("~"), "CONSORCIO CONCREJATOEFFICO LOTE 1", 
                            "Central de Arquivos - QSMS", "000 ATUAL - OBRA 186 - INHAÚMA", 
                            "Logo Consórcio.jpg")

col1, col2, col3 = st.columns([1, 4, 1])  

with col1:
    if os.path.exists(logo_concrejato):
        st.image(logo_concrejato, width=220)
        
    else:
        st.warning("Logotipo não encontrado!")

with col2:
    st.markdown(
        "<h1 style='text-align: center; color: #004080; font-size: 50px;'>AUTOMAÇÃO SERTRAS</h1>", 
        unsafe_allow_html=True
    )  

with col3:
    if os.path.exists(logo_consorcio):
        st.image(logo_consorcio, width=220)
    else:
        st.warning("Logotipo não encontrado!")

# Linha Separadora 
st.markdown("<hr style='border: 1px solid #004080;'>", unsafe_allow_html=True)

placeholder_botao = st.empty()

if not st.session_state["executado"]:
    col_empty1, col_button, col_empty2 = st.columns([2, 2, 2])
    with col_button:
        # 🟢 Botão Azul Nativo do Streamlit
        if st.button("🚀 Executar Automação", key="rodar_automacao", help="Clique para executar a automação", 
                     use_container_width=True, type="primary"):  
            with st.spinner("Executando automação..."):
                tabela, documentos_não_encontrados, documentos_encontrados, documentos_enviados, datas_extraidas, vencimentos_projetados, vencimentos_enviados, erro_envio, documentos_atualizados, documentos_nao_atualizados, datas_modificacao = automacao.run_automation(documentos_rh, documentos_QSMS, diretorio_base_rh, diretorio_base_qsms, mapeamento_para_documentos, mapeamento_para_datas)

                st.session_state["dados_processados"] = {
                    "tabela": tabela,
                    "documentos_não_encontrados": documentos_não_encontrados,
                    "documentos_encontrados": documentos_encontrados,
                    "documentos_nao_atualizados": documentos_nao_atualizados,
                    "documentos_atualizados": documentos_atualizados,
                    "erro_envio": erro_envio,
                    "datas_extraidas": datas_extraidas,
                    "datas_modificacao" : datas_modificacao,
                    "vencimentos_projetados": vencimentos_projetados,
                    "documentos_enviados": documentos_enviados,
                    "vencimentos_enviados": vencimentos_enviados
                }

                st.session_state["executado"] = True
                st.rerun()

if st.session_state["dados_processados"]:
    dados = st.session_state["dados_processados"]
    
    df_sertras = dados["tabela"]

    df_documentos = pd.DataFrame(list(zip_longest(dados["documentos_não_encontrados"], dados["documentos_encontrados"], dados["datas_modificacao"],
                                        dados["documentos_nao_atualizados"], dados["documentos_atualizados"], fillvalue="---")),

                                columns=["DOCUMENTOS NÃO ENCONTRADOS", "DOCUMENTOS ENCONTRADOS", "DATAS MODIFICAÇÃO", "DOCUMENTOS NÃO ATUALIZADOS", "DOCUMENTOS ATUALIZADOS"])

    df_relatorio = pd.DataFrame(list(zip_longest(dados["erro_envio"], dados["datas_extraidas"], dados["vencimentos_projetados"], 
                                       dados["documentos_enviados"], dados["vencimentos_enviados"], fillvalue="---")),

                                columns=["DOCUMENTOS SEM DATA EXTRAÍDA","DATAS EXTRAÍDAS", "VENCIMENTOS PROJETADOS", "DOCUMENTOS ENVIADOS", "VENCIMENTOS ENVIADOS"])
    
    @st.cache_data
    def to_excel_cached(df, sheet_name):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        return output.getvalue()

    excel_sertras = to_excel_cached(df_sertras, "Pendencias_Sertras")
    excel_documentos = to_excel_cached(df_documentos, "Relacao_Documentos")
    excel_relatorio = to_excel_cached(df_relatorio, "Relatorio_Execução")

    centered_style = [
        {"selector": "thead th", "props": [("background-color", "blue"), ("color", "white"), ("font-weight", "bold"), ("text-align", "center")]},
        {"selector": "tbody td", "props": [("text-align", "center")]}
    ]

    df_sertras_html = df_sertras.style.set_table_styles(centered_style).hide(axis="index").to_html()
    df_documentos_html = df_documentos.style.set_table_styles(centered_style).hide(axis="index").to_html()
    df_relatorio_html = df_relatorio.style.set_table_styles(centered_style).hide(axis="index").to_html()

    def exibir_tabela(titulo, df, arquivo_excel, nome_arquivo):
        col1, col2, col3 = st.columns([0.5, 5, 0.5])  
        with col2:
            st.markdown(f"## 📋 {titulo}")
            st.markdown(df,unsafe_allow_html=True)
        with col3:
            st.download_button(
                data=arquivo_excel,
                label="⬇️",
                file_name=nome_arquivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        st.markdown("<br><br>", unsafe_allow_html=True)

    exibir_tabela("RELATÓRIO PENDÊNCIAS SERTRAS", df_sertras_html, excel_sertras, f"PENDÊNCIA_SERTRAS {data_atual}.xlsx")
    exibir_tabela("RELAÇÃO DOCUMENTOS", df_documentos_html, excel_documentos, f"RELAÇÃO_DOCUMENTOS {data_atual}.xlsx")
    exibir_tabela("RELATÓRIO EXECUÇÃO", df_relatorio_html, excel_relatorio, f"RELATÓRIO_EXECUÇÃO {data_atual}.xlsx")


streamlit_rodando = False

for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
    try:
        if proc.info['cmdline'] and any("streamlit" in cmd for cmd in proc.info['cmdline']):
            streamlit_rodando = True
            break  
    except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
        pass

if not streamlit_rodando:
    subprocess.Popen([sys.executable, "-m", "streamlit", "run", "Automação_Completa.py"], shell=True)





        







        

